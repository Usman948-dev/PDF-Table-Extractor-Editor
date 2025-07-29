import streamlit as st
import camelot
import pandas as pd
import io
import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.cell_range import CellRange

# --- Helper Functions and Data ---

# Define a dictionary for unit conversions (extended from previous discussions)
UNIT_CONVERSIONS = {
    "p-no/p hour": 1.0,
    "p-no": 1.0,
    "hour": 1.0,
    "hours": 1.0,
    "meters": 1.0,
    "kilometers": 1000.0,
    "each": 1.0,
    "lot": 1.0,
    "kgs": 1.0,
    "tons": 1000.0,
    "liters": 1.0,
    "gallons": 3.785,
    "days": 8.0,  # Assuming 8 working hours per day for calculation
    "nos": 1.0,  # Common abbreviation for numbers
    "pc": 1.0,  # Pieces
    "set": 1.0,  # Sets
    # Specific percentage units for 'Units No.'
    "%": 100.0,
    "%0": 1000.0,
    "%o": 1000.0,  # Assuming typo for %0
    "sqm": 1.0,  # Square meters
    "cum": 1.0,  # Cubic meters
    "lm": 1.0,  # Linear meters
    # Add more unit mappings as needed: "your unit string": corresponding_number,
}


def convert_unit_to_number(unit_string):
    """
    Converts a unit string to a numerical value based on predefined mappings.
    Prioritizes specific percentage logic if '%', '%0', or '%o' are present,
    otherwise uses the UNIT_CONVERSIONS dictionary.
    """
    if not isinstance(unit_string, str):
        return 0.0  # Return 0 for non-string types (e.g., NaN, None)

    unit_string_cleaned = unit_string.strip().lower()

    # Specific logic for percentage units from the original code
    if "%0" in unit_string_cleaned or "%o" in unit_string_cleaned:
        return 1000.0
    elif "%" in unit_string_cleaned:
        return 100.0

    # Otherwise, use the general UNIT_CONVERSIONS dictionary
    return UNIT_CONVERSIONS.get(unit_string_cleaned, 1.0)  # Default to 1.0 if not found for general units


def calculate_total_rate(input_rate, quantity, units_no):
    """Calculates Total Rate: (Input Rate * Quantity) / Units No."""
    # Convert inputs to numeric, coercing errors to NaN, then fill NaN with 0
    input_rate = pd.to_numeric(input_rate, errors='coerce').fillna(0.0)
    quantity = pd.to_numeric(quantity, errors='coerce').fillna(0.0)
    units_no = pd.to_numeric(units_no, errors='coerce').fillna(1.0)  # Avoid division by zero, treat 0 as 1

    # Ensure units_no is not 0 for division
    units_no = units_no.apply(lambda x: 1.0 if x == 0 else x)

    return (input_rate * quantity) / units_no


def prepare_df_for_editor(df_original, global_estimated_rate_val, pdf_column_mapping_rules, final_excel_column_order):
    """
    Prepares a DataFrame for st.data_editor with all required columns,
    including extracted, calculated, and user-input columns.
    All column names will be flattened strings.
    """
    # 1. Normalize original DataFrame's columns to flat strings
    # This is critical to avoid 'tuple' keys or other non-string column names
    temp_df = df_original.copy()
    temp_df.columns = [str(col).strip() for col in temp_df.columns]  # Ensure all columns are strings

    # Initialize the final DataFrame with all columns from final_excel_column_order
    # This ensures all columns are present from the start
    processed_df = pd.DataFrame(columns=final_excel_column_order)

    # Transfer data from original PDF columns to the new DataFrame
    # This loop now more carefully maps original data to target columns
    for excel_col_name in final_excel_column_order:
        found_in_pdf = False
        # Check if the excel_col_name is a direct mapping from PDF (e.g., 'Sr No.')
        for pdf_target_col, rules in pdf_column_mapping_rules.items():
            if excel_col_name == pdf_target_col:  # Simple direct map
                for keyword in rules["keywords"]:
                    # Try to find a matching column in the *flattened* temp_df columns
                    # Check if the keyword is *in* any original column name (case-insensitive)
                    matching_original_col = next((
                        orig_col for orig_col in temp_df.columns if keyword.lower() in orig_col.lower()
                    ), None)

                    if matching_original_col and matching_original_col in temp_df.columns:
                        # Copy the data to the target column
                        processed_df[excel_col_name] = temp_df[matching_original_col].reset_index(drop=True)
                        found_in_pdf = True
                        break  # Found a match for this excel_col_name's keywords
            if found_in_pdf:
                break

        if not found_in_pdf:
            # If no mapping was found or data wasn't copied, initialize with empty Series
            # Ensure same length as extracted data if applicable, or 0 if temp_df is empty
            series_length = len(temp_df) if not temp_df.empty else 0
            processed_df[excel_col_name] = pd.Series(dtype='object', index=range(
                series_length))  # Ensure same length as extracted data if applicable

    # If the processed_df is empty after mapping, ensure it has at least some rows to handle global input
    if processed_df.empty:
        processed_df = pd.DataFrame(columns=final_excel_column_order, index=[0])  # Create at least one row for inputs

    # Handle 'Estimated Rate - Input' (from PDF or global input)
    # Check if 'Estimated Rate - Input' was extracted from PDF; if not, apply global
    if "Estimated Rate - Input" in processed_df.columns:
        # Check if it's all null/empty strings
        is_all_null = processed_df["Estimated Rate - Input"].apply(lambda x: pd.isna(x) or str(x).strip() == "").all()
        if is_all_null and global_estimated_rate_val:
            try:
                rate_val = float(global_estimated_rate_val)
                processed_df["Estimated Rate - Input"] = rate_val
            except ValueError:
                pass  # Invalid global rate, leave as empty

        # Ensure it's numeric after potential global fill
        processed_df["Estimated Rate - Input"] = pd.to_numeric(processed_df["Estimated Rate - Input"],
                                                               errors='coerce').fillna(0.0)

    # Fill NaN values in 'Quantity', 'Market Rate - Input', 'Quoted Rate - Input' with 0 for calculations
    # This ensures they are numeric before calculations
    for col in ["Quantity", "Market Rate - Input", "Quoted Rate - Input"]:
        if col in processed_df.columns:
            processed_df[col] = pd.to_numeric(processed_df[col], errors='coerce').fillna(0.0)

    # Initialize calculated columns with 0.0 (they will be recalculated by `recalculate_editor_df_values`)
    for col in ["Units No.", "Market Rate - Total", "Estimated Rate - Total", "Quoted Rate - Total"]:
        if col in processed_df.columns:
            processed_df[col] = 0.0

            # Ensure 'Units' column exists for 'Units No.' calculation, initialize with empty string if missing
    if "Units" not in processed_df.columns:
        processed_df["Units"] = ""
    # Ensure 'Sr No.' and 'Items Description' exist and are string type if they are target columns
    if "Sr No." not in processed_df.columns:
        processed_df["Sr No."] = ""
    if "Items Description" not in processed_df.columns:
        processed_df["Items Description"] = ""

    return processed_df


def recalculate_editor_df_values(df):
    """
    Recalculates 'Units No.' and all 'Total Rate' columns
    based on user edits in 'Input Rate', 'Quantity', and 'Unit' columns.
    """
    df_copy = df.copy()  # Work on a copy to avoid SettingWithCopyWarning

    # 1. Recalculate 'Units No.'
    if "Units" in df_copy.columns:
        df_copy["Units No."] = df_copy["Units"].apply(convert_unit_to_number)
    else:
        df_copy["Units No."] = 1.0  # Default if 'Units' column is missing

    # Ensure 'Units No.' is numeric and handle potential zeros for division
    df_copy["Units No."] = pd.to_numeric(df_copy["Units No."], errors='coerce').fillna(1.0)
    # Important: Prevent division by zero if Units No. becomes 0
    df_copy["Units No."] = df_copy["Units No."].apply(lambda x: 1.0 if x == 0 else x)

    # Ensure input rate and quantity columns are numeric
    for col in ["Quantity", "Market Rate - Input", "Estimated Rate - Input", "Quoted Rate - Input"]:
        if col in df_copy.columns:
            df_copy[col] = pd.to_numeric(df_copy[col], errors='coerce').fillna(0.0)

    # 2. Recalculate 'Total Rate' columns
    if "Market Rate - Input" in df_copy.columns and "Quantity" in df_copy.columns and "Units No." in df_copy.columns:
        df_copy["Market Rate - Total"] = calculate_total_rate(
            df_copy["Market Rate - Input"], df_copy["Quantity"], df_copy["Units No."])

    if "Estimated Rate - Input" in df_copy.columns and "Quantity" in df_copy.columns and "Units No." in df_copy.columns:
        df_copy["Estimated Rate - Total"] = calculate_total_rate(
            df_copy["Estimated Rate - Input"], df_copy["Quantity"], df_copy["Units No."])

    if "Quoted Rate - Input" in df_copy.columns and "Quantity" in df_copy.columns and "Units No." in df_copy.columns:
        df_copy["Quoted Rate - Total"] = calculate_total_rate(
            df_copy["Quoted Rate - Input"], df_copy["Quantity"], df_copy["Units No."])

    return df_copy


# --- Streamlit UI ---
st.set_page_config(page_title="PDF Table Extractor and Editor", layout="wide")

st.title("📄 PDF Table Extractor & Editor")
st.markdown("""
Upload a PDF file, specify the pages, and extract tables directly into an interactive editor.
You can then modify the data within the app, and calculated columns will update.

*Important Prerequisites (Install FIRST):*
1.  *Install Ghostscript:* This is an external dependency for camelot.
    * *Windows:* Download installer from [Ghostscript website](https://www.ghostscript.com/download/gsdnld.html) (choose Ghostscript AGPL Release). Install it and **note the full path to the gswin64c.exe file** (e.g., C:\\Program Files\\gs\\gs10.02.1\\bin\\gswin64c.exe).
    * *macOS (Homebrew):* `brew install ghostscript`
    * *Linux (Debian/Ubuntu):* `sudo apt-get install ghostscript`
2.  *Install Python Libraries in your virtual environment:*
    ```bash
    pip install streamlit "camelot-py[cv]" openpyxl pandas
    ```
    (If `camelot-py[cv]` gives issues, try `pip install "camelot-py[all]"`).
""")

# --- Ghostscript Path Input (Persisted) ---
if 'ghostscript_path_value' not in st.session_state:
    st.session_state.ghostscript_path_value = ""

ghostscript_path = st.text_input(
    "Enter full path to Ghostscript executable (e.g., C:\\Program Files\\gs\\gs10.02.1\\bin\\gswin64c.exe). Required for table extraction.",
    value=st.session_state.ghostscript_path_value,
    key="gs_path_input"
)

if st.session_state.ghostscript_path_value != ghostscript_path:
    st.session_state.ghostscript_path_value = ghostscript_path
    st.rerun()

if st.session_state.ghostscript_path_value:
    try:
        camelot.core.postscript_path = st.session_state.ghostscript_path_value
    except Exception as e:
        st.error(f"Could not set Ghostscript path. Please check the path and Ghostscript installation: {e}")
        camelot.core.postscript_path = ''
        st.session_state.ghostscript_path_value = ''
else:
    st.warning("Please provide the path to your Ghostscript executable. Table extraction will not work without it.")

# --- PDF Upload and Extraction Settings ---
uploaded_file = st.file_uploader("Upload your PDF file", type=["pdf"], key="pdf_uploader")

pages_input = st.text_input(
    "Enter pages to extract (e.g., '1,3-5,8'). Leave blank for all pages with tables.",
    value="",
    key="pages_input"
)

flavor_selection = st.radio(
    "Choose Table Extraction Method (Flavor):",
    ('stream', 'lattice'),
    index=1,
    help="""
    *Stream:* Best for tables without ruling lines (columns delimited by whitespace).
    *Lattice:* Best for tables with clearly defined ruling lines. Often more accurate for structured tables.
    """,
    key="flavor_radio"
)

line_scale_input = None
if flavor_selection == 'lattice':
    line_scale_input = st.slider(
        "Adjust Line Scale (for 'lattice' flavor, affects line detection):",
        min_value=1,
        max_value=100,
        value=50,
        help="""
        A higher value makes Camelot detect thinner lines.
        Adjust this if 'lattice' flavor isn't detecting lines correctly.
        """,
        key="line_scale_slider"
    )

# --- Global Inputs for Calculations ---
global_estimated_rate_input = st.text_input(
    "Global Estimated Rate (optional, e.g., 25.50). This will pre-fill 'Estimated Rate - Input' if extracted is empty.",
    value="",
    key="global_estimated_rate_input"
)

# --- Formula Type Selection (affects displayed columns and download) ---
formula_type = st.selectbox(
    "Select Calculation Type for Total Column:",
    ("Default (Estimated Rate)", "Quoted Rate Calculation"),
    help="""
    *Default (Estimated Rate):* Uses 'Estimated Rate - Input' to calculate 'Estimated Rate - Total'. No 'Quoted Rate' option.
    *Quoted Rate Calculation:* Shows 'Estimated Rate - Input'/'Total', and adds 'Quoted Rate - Input'/'Total' for comparison.
    """,
    key="formula_type_select"
)

# --- Name Input for Excel Sheet (also used for summary) ---
name_input = st.text_input("Name (will appear in summary and Excel sheet)", value="Name:", key="name_input")

# --- Extraction Button ---
extract_button = st.button("Extract Tables for Editing", key="extract_button")

# --- Define PDF Column Mapping Rules ---
pdf_column_mapping_rules = {
    "Sr No.": {"keywords": ["sr.no", "sr no.", "sr no", "seriel number", "srno", "s.no", "serial no", "serial number",
                            "sr. no.", "sr.", "no."], "include_header_cell_in_data": False},
    "Items Description": {
        "keywords": ["items", "item description", "items description", "item name", "items name", "item", "description",
                     "item desc"], "include_header_cell_in_data": False},
    "Units": {"keywords": ["units", "unit"], "include_header_cell_in_data": False},
    "Quantity": {
        "keywords": ["quantity", "quantities", "estimated quantity", "estimated qty", "qty", "est. qty", "est qty",
                     "estimated quanity"], "include_header_cell_in_data": False},
    # Mapping for 'Estimated Rate - Input'
    "Estimated Rate - Input": {
        "keywords": ["est", "est rates", "est rate", "estimated rates", "estimated rate", "rate", "rates"],
        "include_header_cell_in_data": False},
    # Mapping for 'Market Rate - Input' (new)
    "Market Rate - Input": {"keywords": ["market rate", "mkt rate", "mrkt rate", "market rates"],
                            "include_header_cell_in_data": False},
}

# --- Define Final Column Order for Editor and Excel ---
# This list dictates the order of columns in the displayed data editor AND the final Excel output.
base_columns = ["Sr No.", "Items Description", "Units", "Units No.", "Quantity"]

if formula_type == "Default (Estimated Rate)":
    final_display_excel_column_order = base_columns + [
        "Estimated Rate - Input",
        "Estimated Rate - Total"
    ]
elif formula_type == "Quoted Rate Calculation":
    final_display_excel_column_order = base_columns + [
        "Estimated Rate - Input",
        "Estimated Rate - Total",
        "Market Rate - Input",  # Added Market Rate columns
        "Market Rate - Total",
        "Quoted Rate - Input",
        "Quoted Rate - Total"
    ]

# --- Main Extraction and Display Logic ---
if extract_button:
    if not st.session_state.ghostscript_path_value:
        st.error("Please provide the path to your Ghostscript executable.")
    elif not uploaded_file:
        st.warning("Please upload a PDF file first!")
    else:
        pdf_stream = io.BytesIO(uploaded_file.read())

        try:
            pages_arg = 'all' if pages_input.strip() == "" else pages_input

            with st.spinner(f"Extracting tables using '{flavor_selection}' method..."):
                read_pdf_kwargs = {
                    'pages': pages_arg,
                    'flavor': flavor_selection,
                    'split_text': True,
                }
                if flavor_selection == 'lattice' and line_scale_input is not None:
                    read_pdf_kwargs['line_scale'] = line_scale_input

                tables = camelot.read_pdf(
                    pdf_stream,
                    **read_pdf_kwargs
                )

            if len(tables) == 0:
                st.warning("No tables were found on the specified pages.")
                st.info("Try adjusting page numbers, selection method, or line scale.")
                st.stop()
            else:
                st.success(f"Successfully extracted {len(tables)} table(s).")
                sorted_tables = sorted(tables, key=lambda t: t.page)

                # --- Identify Headers ONLY from the FIRST table ---
                first_table_obj = sorted_tables[0]
                df_first_page = first_table_obj.df.copy()
                # Ensure columns are flattened immediately for header detection
                df_first_page.columns = [str(col).strip() for col in df_first_page.columns]

                best_header_row_idx_first_page = -1
                max_matches_first_page = 0
                header_column_map_from_first_page = {}  # {target_col_name: original_camelot_col_index}

                # Iterate through potential header rows (up to first 5 rows)
                for row_idx in range(min(5, df_first_page.shape[0])):
                    current_row_values_cleaned = [str(cell).strip().lower() for cell in
                                                  df_first_page.iloc[row_idx].tolist()]
                    current_matches = 0
                    temp_header_col_map = {}

                    for target_col_name, rules in pdf_column_mapping_rules.items():
                        # Look for any keyword in the header row cells
                        for i, header_cell_content in enumerate(current_row_values_cleaned):
                            if any(keyword in header_cell_content for keyword in rules["keywords"]):
                                temp_header_col_map[target_col_name] = i
                                current_matches += 1
                                break

                    if current_matches > max_matches_first_page:
                        max_matches_first_page = current_matches
                        best_header_row_idx_first_page = row_idx
                        header_column_map_from_first_page = temp_header_col_map.copy()

                    if max_matches_first_page == len(pdf_column_mapping_rules):  # All required columns found
                        break

                all_processed_dfs_for_concat = []  # List to hold all processed DataFrames before concat

                if best_header_row_idx_first_page == -1:
                    st.info("Could not identify a clear header row with key columns. Displaying raw extraction.")
                    # Fallback: Process all tables as is, without header row removal, and rely on user to clean
                    for table_obj in sorted_tables:
                        df_raw = table_obj.df.copy()
                        # Flatten columns just in case
                        df_raw.columns = [str(col).strip() for col in df_raw.columns]

                        # Create a basic df for editor with all expected columns
                        editor_df = pd.DataFrame(columns=final_display_excel_column_order)

                        # Copy data from raw_df to editor_df based on column names
                        for col in final_display_excel_column_order:
                            if col in df_raw.columns:  # If a column with this exact name exists in raw data
                                editor_df[col] = df_raw[col].reset_index(drop=True)
                            else:
                                editor_df[col] = pd.Series(dtype='object',
                                                           index=range(len(df_raw)))  # Ensure same length

                        # Apply initial calculations (Units No., Totals)
                        editor_df = recalculate_editor_df_values(editor_df)
                        all_processed_dfs_for_concat.append(editor_df)  # Add to list for concat

                    st.session_state.header_detection_failed = True  # Flag
                else:
                    st.success(
                        f"Headers identified on page {first_table_obj.page}, row {best_header_row_idx_first_page + 1}. Applying this structure.")
                    st.session_state.header_detection_failed = False

                    for table_obj in sorted_tables:
                        df = table_obj.df.copy()
                        # Flatten columns immediately after copying for consistency
                        df.columns = [str(col).strip() for col in df.columns]

                        current_processed_df_data = {}  # To hold data for new df

                        # Collect data from relevant columns based on header mapping
                        start_data_row = best_header_row_idx_first_page + 1 if table_obj.page == first_table_obj.page else 0  # Start from 0 for subsequent pages

                        # Iterate through the expected columns and map data
                        for target_col_name in final_display_excel_column_order:
                            # Check if this target_col_name was found in the PDF header mapping
                            original_col_index = header_column_map_from_first_page.get(target_col_name)

                            if original_col_index is not None and original_col_index < df.shape[1]:
                                col_data = df.iloc[start_data_row:, original_col_index].reset_index(drop=True)
                                current_processed_df_data[target_col_name] = col_data
                            else:
                                # If not found in PDF, or if index is out of bounds, initialize as empty
                                current_processed_df_data[target_col_name] = pd.Series(dtype='object', index=range(
                                    df.shape[0] - start_data_row))

                        # Create a DataFrame from the collected data, ensuring index alignment
                        # pd.DataFrame will align by column names automatically
                        extracted_flat_df = pd.DataFrame(current_processed_df_data)

                        # Now, prepare this DataFrame for the editor, including all target columns and calculations
                        editor_df = prepare_df_for_editor(
                            extracted_flat_df,
                            global_estimated_rate_input,
                            pdf_column_mapping_rules,
                            final_display_excel_column_order
                        )

                        # Perform initial calculation on the prepared DF
                        editor_df = recalculate_editor_df_values(editor_df)

                        all_processed_dfs_for_concat.append(editor_df)  # Add to list for concat

                # CONCATENATE ALL PROCESSED TABLES INTO A SINGLE DATAFRAME
                if all_processed_dfs_for_concat:
                    st.session_state.single_combined_df = pd.concat(all_processed_dfs_for_concat, ignore_index=True)
                    st.session_state.extraction_success = True
                else:
                    st.info("No tables could be processed with the specified column headers or no tables found at all.")
                    st.session_state.extraction_success = False  # No data to display

        except Exception as e:
            st.error(f"An error occurred during PDF extraction: {e}")
            st.info(
                "Please check your PDF file, Ghostscript installation, or page number input. Error details: " + str(e))
            st.session_state.extraction_success = False  # Indicate failure

# --- Display Data Editor if Extraction was Successful and data exists ---
if 'extraction_success' in st.session_state and st.session_state.extraction_success and 'single_combined_df' in st.session_state and not st.session_state.single_combined_df.empty:
    st.subheader("Edit Extracted Tables")
    st.caption("Double-click a cell to edit. Press 'Enter' or click outside to see calculations update.")
    st.caption("Grayed out columns are calculated and not editable directly.")

    # Get the current combined DataFrame from session state
    df_for_editor = st.session_state.single_combined_df.copy()  # Work on a copy for the widget

    # Define column configurations for st.data_editor
    column_config_dict = {
        "Units No.": st.column_config.NumberColumn("Units No.", help="Calculated from Units", format="%.2f",
                                                   disabled=True),
        "Market Rate - Total": st.column_config.NumberColumn("Market Rate - Total",
                                                             help="Calculated from Market Rate - Input * Quantity / Units No.",
                                                             format="$%.2f", disabled=True),
        "Estimated Rate - Total": st.column_config.NumberColumn("Estimated Rate - Total",
                                                                help="Calculated from Estimated Rate - Input * Quantity / Units No.",
                                                                format="$%.2f", disabled=True),
        "Quoted Rate - Total": st.column_config.NumberColumn("Quoted Rate - Total",
                                                             help="Calculated from Quoted Rate - Input * Quantity / Units No.",
                                                             format="$%.2f", disabled=True),

        "Units": st.column_config.TextColumn("Units", help="Enter unit (e.g., 'P-no/P Hour'). Affects 'Units No.'.",
                                             width="small"),
        "Quantity": st.column_config.NumberColumn("Quantity", help="Number of items or amount.", format="%.2f"),
        "Market Rate - Input": st.column_config.NumberColumn("Market Rate - Input", help="Enter the market input rate.",
                                                             format="$%.2f"),
        "Estimated Rate - Input": st.column_config.NumberColumn("Estimated Rate - Input",
                                                                help="Enter the estimated input rate (can be pre-filled).",
                                                                format="$%.2f"),
        "Quoted Rate - Input": st.column_config.NumberColumn("Quoted Rate - Input",
                                                             help="Enter the quoted input rate (for comparison).",
                                                             format="$%.2f"),

        # Default behavior for other columns: text, non-editable
        "Sr No.": st.column_config.TextColumn("Sr No.", disabled=True),
        "Items Description": st.column_config.TextColumn("Items Description", disabled=True, width="large"),
    }

    # Ensure that only columns actually present in the final_display_excel_column_order are configured
    # And set default disabled/enabled status for other columns
    active_column_config = {}
    editable_cols = ["Units", "Quantity", "Market Rate - Input", "Estimated Rate - Input", "Quoted Rate - Input"]

    for col_name in final_display_excel_column_order:
        if col_name in df_for_editor.columns:  # Only configure if column exists in the current dataframe
            if col_name in column_config_dict:
                active_column_config[col_name] = column_config_dict[col_name]
            elif col_name in editable_cols:
                active_column_config[col_name] = st.column_config.TextColumn(label=col_name, disabled=False)
            else:
                active_column_config[col_name] = st.column_config.TextColumn(label=col_name, disabled=True)

    edited_df_from_widget = st.data_editor(
        df_for_editor,  # Display the current combined DataFrame
        key="single_combined_data_editor",  # Single key for the combined editor
        hide_index=True,
        num_rows="dynamic",
        column_config=active_column_config
    )

    # Recalculate based on the *edited* DataFrame from the widget
    recalculated_combined_df = recalculate_editor_df_values(edited_df_from_widget)

    # Update the session state with the latest recalculated combined DataFrame
    st.session_state.single_combined_df = recalculated_combined_df

    st.markdown("---")

    # --- Summary Calculations & Display ---
    if not st.session_state.single_combined_df.empty:
        combined_final_df = st.session_state.single_combined_df  # Use the latest from session state

        st.subheader("Summary")
        st.write(f"**Name:** {name_input}")

        total_estimated_rate_sum = combined_final_df[
            "Estimated Rate - Total"].sum() if "Estimated Rate - Total" in combined_final_df.columns else 0
        total_quoted_rate_sum = combined_final_df[
            "Quoted Rate - Total"].sum() if "Quoted Rate - Total" in combined_final_df.columns else 0

        st.metric(label="Grand Total (Estimated)", value=f"${total_estimated_rate_sum:,.2f}")

        if formula_type == "Quoted Rate Calculation":
            st.metric(label="Grand Total (Quoted)", value=f"${total_quoted_rate_sum:,.2f}")

            if total_estimated_rate_sum > 0:
                rate_below_govt = 1 - (total_quoted_rate_sum / total_estimated_rate_sum)
                st.metric(label="Rate Below From Govt Rate", value=f"{rate_below_govt:.2%}")
            else:
                st.info("Cannot calculate 'Rate Below From Govt Rate': Estimated Grand Total is zero.")

        st.markdown("---")
        st.success("You can now download the combined, edited data as Excel.")

        # --- Excel Download Button ---
        excel_output_buffer = io.BytesIO()
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create a single sheet for the combined data
        sheet_name = "Combined Tables"
        ws = wb.create_sheet(title=sheet_name)

        col_to_excel_letter = {col_name: get_column_letter(idx + 1)
                               for idx, col_name in enumerate(final_display_excel_column_order)}

        initial_data_row = 4  # Start data and headers from row 4

        # Reserve top rows for summary formulas (if applicable) and name
        for _ in range(initial_data_row - 1):
            ws.append([])

        # Write headers
        ws.append(final_display_excel_column_order)

        # Apply styling to header row
        header_row_obj = ws[initial_data_row]
        for cell in header_row_obj:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')

        # Freeze the header row
        ws.freeze_panes = get_column_letter(1) + str(initial_data_row + 1)

        # Write data rows and apply formulas
        for r_idx, row_data_series in combined_final_df.iterrows():  # Loop through the single combined_final_df
            excel_row_num = r_idx + initial_data_row + 1
            row_values_for_excel = []
            for col_name in final_display_excel_column_order:
                if col_name in ["Market Rate - Total", "Estimated Rate - Total", "Quoted Rate - Total"]:
                    # Construct Excel formula for Total columns
                    input_rate_col_name = ""
                    if col_name == "Market Rate - Total":
                        input_rate_col_name = "Market Rate - Input"
                    elif col_name == "Estimated Rate - Total":
                        input_rate_col_name = "Estimated Rate - Input"
                    elif col_name == "Quoted Rate - Total":
                        input_rate_col_name = "Quoted Rate - Input"

                    input_rate_col_letter = col_to_excel_letter.get(input_rate_col_name)
                    quantity_col_letter = col_to_excel_letter.get("Quantity")
                    units_no_col_letter = col_to_excel_letter.get("Units No.")

                    if all(c in final_display_excel_column_order for c in
                           [input_rate_col_name, "Quantity", "Units No."]) and \
                            input_rate_col_letter and quantity_col_letter and units_no_col_letter:
                        # Using IFERROR and VALUE to handle potential non-numeric inputs in Excel gracefully
                        formula_str = (
                            f"=(IFERROR(VALUE({input_rate_col_letter}{excel_row_num}),0)*IFERROR(VALUE({quantity_col_letter}{excel_row_num}),0))/"
                            f"(IF(OR(ISBLANK({units_no_col_letter}{excel_row_num}),IFERROR(VALUE({units_no_col_letter}{excel_row_num}),1)=0),1,IFERROR(VALUE({units_no_col_letter}{excel_row_num}),1)))"
                        )
                        row_values_for_excel.append(formula_str)
                    else:
                        row_values_for_excel.append(None)  # Fallback if columns not found
                else:
                    cell_value = row_data_series.get(col_name)
                    # Ensure non-numeric NaNs are written as empty strings or None for Excel
                    row_values_for_excel.append(None if pd.isna(cell_value) else str(cell_value).strip())
            ws.append(row_values_for_excel)

        # --- Styling and Borders for the single sheet ---
        for col_idx, column_header in enumerate(final_display_excel_column_order):
            col_letter = get_column_letter(col_idx + 1)
            if column_header == "Items Description":
                ws.column_dimensions[col_letter].width = 70
                for r in range(initial_data_row + 1, ws.max_row + 1):
                    ws.cell(row=r, column=col_idx + 1).alignment = Alignment(wrap_text=True, vertical='top')
            else:
                ws.column_dimensions[col_letter].width = 15  # Slightly wider default for better number display

        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 18

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=initial_data_row, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border

        # --- Grand Total Rows for the single sheet ---
        current_data_end_row = ws.max_row

        # Grand Total for "Estimated Rate - Total"
        est_total_col_name = "Estimated Rate - Total"
        if est_total_col_name in final_display_excel_column_order:
            est_total_col_letter = col_to_excel_letter.get(est_total_col_name)
            if est_total_col_letter:
                start_row_for_sum = initial_data_row + 1
                end_row_for_sum = current_data_end_row

                grand_total_label_row_est = [None] * len(final_display_excel_column_order)
                if "Sr No." in final_display_excel_column_order:
                    grand_total_label_row_est[
                        final_display_excel_column_order.index("Sr No.")] = "Grand Total (Estimated)"
                ws.append(grand_total_label_row_est)
                row_num_for_estimated_grand_total = ws.max_row

                grand_total_formula_cell_est = ws.cell(row=row_num_for_estimated_grand_total,
                                                       column=final_display_excel_column_order.index(
                                                           est_total_col_name) + 1)
                grand_total_formula_cell_est.value = f"=SUM({est_total_col_letter}{start_row_for_sum}:{est_total_col_letter}{end_row_for_sum})"
                for cell in ws[row_num_for_estimated_grand_total]:
                    cell.font = Font(bold=True)
                    cell.border = thin_border

        # Grand Total for "Quoted Rate - Total" (if applicable)
        if formula_type == "Quoted Rate Calculation":
            quoted_total_col_name = "Quoted Rate - Total"
            if quoted_total_col_name in final_display_excel_column_order:
                quoted_total_col_letter = col_to_excel_letter.get(quoted_total_col_name)
                if quoted_total_col_letter:
                    start_row_for_sum = initial_data_row + 1
                    end_row_for_sum = current_data_end_row

                    grand_total_label_row_quoted = [None] * len(final_display_excel_column_order)
                    if "Sr No." in final_display_excel_column_order:
                        grand_total_label_row_quoted[
                            final_display_excel_column_order.index("Sr No.")] = "Grand Total (Quoted)"
                    ws.append(grand_total_label_row_quoted)
                    row_num_for_quoted_grand_total = ws.max_row

                    grand_total_formula_cell_quoted = ws.cell(row=row_num_for_quoted_grand_total,
                                                              column=final_display_excel_column_order.index(
                                                                  quoted_total_col_name) + 1)
                    grand_total_formula_cell_quoted.value = f"=SUM({quoted_total_col_letter}{start_row_for_sum}:{quoted_total_col_letter}{end_row_for_sum})"
                    for cell in ws[row_num_for_quoted_grand_total]:
                        cell.font = Font(bold=True)
                        cell.border = thin_border

        # --- Place "Name" and "Rate Below From Govt Rate" summary on top ---
        if name_input:
            ws['A1'] = name_input
            ws['A1'].font = Font(bold=True)
            ws['A1'].border = thin_border

        if formula_type == "Quoted Rate Calculation":
            est_rate_total_col_name = "Estimated Rate - Total"
            quoted_rate_total_col_name = "Quoted Rate - Total"

            est_rate_total_col_letter = col_to_excel_letter.get(est_rate_total_col_name)
            quoted_rate_total_col_letter = col_to_excel_letter.get(quoted_total_col_name)

            if est_rate_total_col_letter and quoted_rate_total_col_letter:
                ws[f'{est_rate_total_col_letter}1'] = 'Rate Below From Govt Rate'  # Place label
                ws[f'{est_rate_total_col_letter}1'].font = Font(bold=True)
                ws[f'{est_rate_total_col_letter}1'].alignment = Alignment(horizontal='right')
                ws[f'{est_rate_total_col_letter}1'].border = thin_border
                ws[f'{est_rate_total_col_letter}1'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC",
                                                                       fill_type="solid")

                formula_cell_loc = f'{quoted_rate_total_col_letter}1'

                # Use sheet-specific grand total cell references for the formula
                # Ensure row_num_for_estimated_grand_total and row_num_for_quoted_grand_total are defined
                # These are defined within the sheet loop, so they should be available.
                if 'row_num_for_estimated_grand_total' in locals() and 'row_num_for_quoted_grand_total' in locals():
                    formula_summary_str = (
                        f'=IFERROR(1-(IFERROR(VALUE({quoted_rate_total_col_letter}{row_num_for_quoted_grand_total}),0)/'
                        f'IF(IFERROR(VALUE({est_rate_total_col_letter}{row_num_for_estimated_grand_total}),0)=0,1,IFERROR(VALUE({est_rate_total_col_letter}{row_num_for_estimated_grand_total}),0))),0)'
                    )
                    ws[formula_cell_loc] = formula_summary_str
                    ws[formula_cell_loc].number_format = '0.00%'
                    ws[formula_cell_loc].font = Font(bold=True)
                    ws[formula_cell_loc].border = thin_border
                    ws[formula_cell_loc].fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
                else:
                    ws[formula_cell_loc] = "N/A"  # Fallback if grand total rows not found
            else:
                st.warning("Could not find required columns for 'Rate Below From Govt Rate' summary.")

        wb.save(excel_output_buffer)
        excel_output_buffer.seek(0)

        st.download_button(
            label="Download Edited Tables as Excel (.xlsx)",
            data=excel_output_buffer.getvalue(),
            file_name="edited_tables_with_formulas.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("No tables to display or download after processing.")

st.markdown("---")
st.markdown("Developed with ❤ using Streamlit, Camelot, Pandas, OpenPyXL.")